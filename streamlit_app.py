import streamlit as st
import pandas as pd
import re
import io
import time
import json # 用于处理 Lottie 动画的JSON文件
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

# --- Lottie 动画加载函数 ---
def load_lottieurl(url: str):
    import requests
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

# Lottie 动画 URLs (可以根据需要替换)
# 'loading' 动画：https://assets3.lottiefiles.com/packages/lf20_kdqph9hh.json
# 'success' 动画：https://assets3.lottiefiles.com/packages/lf20_wkoxp75e.json
# 'error' 动画：https://assets3.lottiefiles.com/packages/lf20_jgK9N4.json
lottie_loading = load_lottieurl("https://assets3.lottiefiles.com/packages/lf20_kdqph9hh.json")
lottie_success = load_lottieurl("https://assets3.lottiefiles.com/packages/lf20_wkoxp75e.json")
lottie_error = load_lottieurl("https://assets3.lottiefiles.com/packages/lf20_jgK9N4.json")

# --- 1. 网页样式美化 (CSS) ---
st.set_page_config(page_title="智能汇总大师", page_icon="✨", layout="centered")

st.markdown("""
    <style>
    /* 渐变背景 */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    /* 标题特效 */
    .main-title {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        color: #1e3a8a;
        text-align: center;
        font-weight: 800;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 20px;
    }
    /* 按钮样式 */
    div.stButton > button:first-child {
        background-color: #4facfe;
        color: white;
        border-radius: 10px;
        border: none;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        font-size: 1.1em;
        padding: 10px 20px;
    }
    div.stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(79, 172, 254, 0.4);
        background-image: linear-gradient(120deg, #4facfe 0%, #00f2fe 100%);
    }
    /* 文件上传区域美化 */
    .stFileUploader > div {
        border: 2px dashed #4facfe;
        border-radius: 10px;
        padding: 20px;
        text-align: center;
        background-color: rgba(255,255,255,0.7);
        transition: all 0.3s ease;
    }
    .stFileUploader > div:hover {
        background-color: rgba(255,255,255,0.9);
        box-shadow: 0 0 10px rgba(79, 172, 254, 0.3);
    }
    /* 提示信息 */
    .stAlert {
        border-radius: 10px;
    }
    </style>
    """, unsafe_allow_html=True)

# 确保 Lottie 动画库已安装
try:
    from streamlit_lottie import st_lottie
except ImportError:
    st.warning("安装 streamlit_lottie 库以显示动画: pip install streamlit-lottie")
    st_lottie = None

# --- 2. 界面展示 ---
st.markdown("<h1 class='main-title'>✨ 智能商品属性汇总大师 ✨</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; color: #64748b; font-size: 1.1em;'>一键上传，精准解析，轻松获取美化报表！</p>", unsafe_allow_html=True)

with st.expander("💡 点击查看详细操作指南"):
    st.markdown("""
    - **第一步**: 点击下方区域上传您的原始 Excel 文件 (`.xlsx` 格式)。
    - **第二步**: 系统将自动解析 G 列（SKU 属性）中的颜色和尺码信息。
    - **第三步**: 解析完成后，屏幕会出现下载按钮，点击即可获取您美化后的汇总报表。
    - **文件格式**: 请确保 G 列包含 'Color' 和 'Size' 关键词，I 列为购买数量。
    """)

# --- 3. 文件上传区域 ---
uploaded_file = st.file_uploader(" ", type=["xlsx"], help="请上传包含商品属性的Excel文件") # 标签设为空，让样式更统一

if uploaded_file:
    # --- 加载动画 ---
    if st_lottie and lottie_loading:
        st_lottie(lottie_loading, speed=1, width=150, height=150, key="loading_anim")
    else:
        st.info("🚀 正在努力解析中，请稍候...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()

    try:
        status_text.text("🔍 正在读取 Excel 数据...")
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        progress_bar.progress(10)
        
        # --- 解析逻辑 (保持你的核心规则不变) ---
        COLOR_REG = r'(?i)Color[:：\s]*([a-zA-Z0-9\-_/]+)'
        SIZE_REG = r'(?i)Size[:：\s]*([a-zA-Z0-9\-\s/]+?)(?=\s*(?:Color|Size|$|[,，;；]))'
        SIZE_MAP = {'HIGH ANKLE SOCKS': 'L', 'KNEE-HIGH SOCKS': 'M'}
        BLUE_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        col_a, col_c, col_g, col_i = df.columns[0], df.columns[2], df.columns[6], df.columns[8]
        all_normal_data, all_error_rows = [], [] # 增加错误行的收集

        status_text.text("⚡ 正在逐行解析商品属性...")
        for index, row in df.iterrows():
            c_raw = str(row[col_c]).strip()
            if not c_raw or c_raw == 'nan': continue
            
            category_name = c_raw.split(' ')[0].upper()
            if category_name.startswith('WZ'): category_name = 'WZ'

            i_nums = re.findall(r'\d+', str(row[col_i]))
            i_qty = int(i_nums[0]) if i_nums else 0
            
            g_text = str(row[col_g])
            chunks = re.split(r'[;；,，\n]', g_text)
            data_pairs = []

            for chunk in chunks:
                chunk = chunk.strip()
                if not chunk: continue
                c_match = re.search(COLOR_REG, chunk)
                s_match = re.search(SIZE_REG, chunk)
                if c_match:
                    color_val = c_match.group(1).strip().upper()
                    raw_size = s_match.group(1).strip().upper() if s_match else ""
                    data_pairs.append((color_val, SIZE_MAP.get(raw_size, raw_size)))
            
            if len(data_pairs) == i_qty and i_qty > 0:
                for cv, sv in data_pairs:
                    all_normal_data.append({'Category': category_name, 'Color': cv, 'Size': sv})
            else:
                all_error_rows.append({'商品名称': category_name, '订单编号': row[col_a], 'SKU属性': g_text}) # 简化错误记录
            
            progress_bar.progress(min(int((index + 1) / len(df) * 90), 90)) # 更新进度条

        status_text.text("🎨 正在生成并美化报表...")
        progress_bar.progress(95)

        # --- 生成内存中的 Excel 文件供下载 ---
        output = io.BytesIO()
        if all_normal_data:
            final_df = pd.DataFrame(all_normal_data)
            categories = sorted(final_df['Category'].unique())
            size_order = ['XXS', 'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', 'FREE', '']
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                output_rows, category_blocks = [], []
                for cat in categories:
                    start_row = len(output_rows) + 1
                    cat_data = final_df[final_df['Category'] == cat]
                    distinct_sizes = sorted(cat_data['Size'].unique(), key=lambda x: size_order.index(x) if x in size_order else 99)
                    output_rows.append({'A': cat})
                    colors = sorted(cat_data['Color'].unique(), key=lambda x: int(re.findall(r'\d+', str(x))[0]) if re.findall(r'\d+', str(x)) else 999)
                    for color in colors:
                        c_data = cat_data[cat_data['Color'] == color]
                        counts = c_data['Size'].value_counts()
                        row_dict = {'A': f"Color {color}"}
                        for idx, s_name in enumerate(distinct_sizes):
                            col_key = chr(66 + idx) if idx < 25 else f"Z{idx}"
                            if s_name in counts:
                                row_dict[col_key] = f"*{counts[s_name]}" if s_name == "" else f"{s_name}*{counts[s_name]}"
                        output_rows.append(row_dict)
                    category_blocks.append((start_row, len(output_rows), 1 + len(distinct_sizes)))
                    output_rows.append({})
                
                pd.DataFrame(output_rows).to_excel(writer, index=False, header=False, sheet_name='汇总结果')
                
                # 美化
                ws = writer.sheets['汇总结果']
                for start, end, col_limit in category_blocks:
                    ws.cell(row=start, column=1).alignment = Alignment(horizontal='center')
                    for r in range(start + 1, end + 1):
                        for c in range(1, col_limit + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.fill, cell.border, cell.alignment = BLUE_FILL, THIN_BORDER, Alignment(horizontal='center')
                for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 15

            progress_bar.progress(100)
            status_text.text("✨ 报表已生成！")
            
            # --- 处理成功后的动画和下载按钮 ---
            st.success("🎉 数据解析成功！您的汇总表已准备就绪。")
            st.balloons() # 庆祝动画
            if st_lottie and lottie_success:
                st_lottie(lottie_success, speed=1, width=100, height=100, key="success_anim")

            st.download_button(
                label="📥 点击下载汇总表",
                data=output.getvalue(),
                file_name=f"汇总_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            # 错误信息显示
            if all_error_rows:
                st.warning(f"⚠️ 发现 {len(all_error_rows)} 行数据未能完全匹配。详细信息请查看下方。")
                st.dataframe(pd.DataFrame(all_error_rows)) # 展示错误行
        else:
            st.error("⚠️ 未能识别到任何有效数据，请检查 G 列内容格式或 Excel 文件是否为空。")
            if st_lottie and lottie_error:
                st_lottie(lottie_error, speed=1, width=100, height=100, key="error_anim")

    except Exception as e:
        st.error(f"❌ 程序运行出错: {e}")
        if st_lottie and lottie_error:
            st_lottie(lottie_error, speed=1, width=100, height=100, key="error_anim_exception")
    finally:
        progress_bar.empty() # 清除进度条
        status_text.empty() # 清除状态文本

st.markdown("---")
st.caption("Powered by Streamlit ✨ | 智能数据处理，让工作更简单。")
